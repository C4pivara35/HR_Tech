import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create Master Workbook with 3 Sheets
wb = openpyxl.Workbook()

# Colors
COLOR_HEADER_BG = "0F172A"       # Dark Slate
COLOR_HEADER_FG = "FFFFFF"       # White
COLOR_ACCENT = "0D9488"          # Teal
COLOR_ZEBRA = "F8FAFC"           # Light slate
COLOR_SUBTOTAL_BG = "F0FDF4"     # Light green
COLOR_SUBTOTAL_FG = "166534"     # Dark green
COLOR_BORDER = "CBD5E1"          # Gray border

font_title = Font(name="Arial", size=15, bold=True, color="0F172A")
font_subtitle = Font(name="Arial", size=11, bold=True, color="0D9488")
font_sec = Font(name="Arial", size=12, bold=True, color="0F172A")
font_th = Font(name="Arial", size=10, bold=True, color=COLOR_HEADER_FG)
font_td = Font(name="Arial", size=10, color="1E293B")
font_td_bold = Font(name="Arial", size=10, bold=True, color="1E293B")
font_subtotal = Font(name="Arial", size=10, bold=True, color=COLOR_SUBTOTAL_FG)
font_total_label = Font(name="Arial", size=11, bold=True, color="0F172A")
font_total_val = Font(name="Arial", size=11, bold=True, color=COLOR_SUBTOTAL_FG)
font_badge = Font(name="Arial", size=13, bold=True, color="FFFFFF")

fill_th = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
fill_subtotal = PatternFill(start_color=COLOR_SUBTOTAL_BG, end_color=COLOR_SUBTOTAL_BG, fill_type="solid")
fill_accent = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# SHEET 1: TCF
ws1 = wb.active
ws1.title = "1. Fatores Técnicos (TCF)"
ws1.views.sheetView[0].showGridLines = True

ws1['A1'] = "AVALIAÇÃO DO FATOR DE COMPLEXIDADE TÉCNICA (TCF)"
ws1['A1'].font = font_title
ws1['A2'] = "Método de Pontos por Casos de Uso (Karner, 1993) — HRTech Core"
ws1['A2'].font = font_subtitle

tcf_headers = ["ID", "Descrição do Fator", "Peso", "Nota (0-5)", "Subtotal (Peso × Nota)", "Justificativa Técnica no Contexto do Projeto"]
ws1.row_dimensions[4].height = 26

for col_num, h in enumerate(tcf_headers, 1):
    c = ws1.cell(row=4, column=col_num, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

tcf_data = [
    ("T1", "Sistemas distribuídos", 2.0, 4, "Arquitetura Web SPA cliente-servidor distribuída com APIs REST e multi-tenancy em nuvem."),
    ("T2", "Objetivos de desempenho e tempo de resposta", 1.0, 4, "Crítico para registros de ponto em horários de pico (resposta esperada < 200 ms — RNF02)."),
    ("T3", "Eficiência do usuário final (on-line)", 1.0, 4, "Navegação SPA fluida, painéis resumidos e aprovações em 1 clique para agilizar a operação."),
    ("T4", "Complexidade interna de processamento", 1.0, 3, "Algoritmos de cálculo de saldo de banco de horas, escalas 12x36, abonos e dupla custódia."),
    ("T5", "Código deve ser reutilizado", 1.0, 5, "Essencial (LPS). Meta de reuso > 75% com 6 ativos reutilizáveis documentados (AR01 a AR06)."),
    ("T6", "Facilidade de instalação", 0.5, 3, "SaaS em nuvem sem instalação no cliente, mas com setup automático de novos tenants em < 8h."),
    ("T7", "Facilidade de uso (Usabilidade)", 0.5, 4, "Atende desde diretores de RH até operadores de fábrica, exigindo alta usabilidade e clareza."),
    ("T8", "Portabilidade", 2.0, 4, "Suporte a múltiplos navegadores em desktops, tablets industriais e smartphones."),
    ("T9", "Facilidade de mudança (Manutenibilidade)", 1.0, 5, "Essencial. Arquitetura modular desacoplada para ativação de regras e módulos por tenant."),
    ("T10", "Concorrência", 1.0, 4, "Elevado número de acessos simultâneos ao bater ponto nas trocas de turno e ao consultar dashboards."),
    ("T11", "Características especiais de segurança", 1.0, 4, "Controle RBAC, trilha de auditoria imutável, conformidade com a LGPD e dupla custódia."),
    ("T12", "Acesso direto para terceiros", 1.0, 3, "Integração API com cartões de benefícios (Caju/Flash), eSocial e Portal do Corretor parceiro."),
    ("T13", "Treinamento especial dos usuários", 1.0, 2, "Interface autoexplicativa com design intuitivo para dispensar treinamentos complexos.")
]

for idx, (fid, desc, weight, score, justif) in enumerate(tcf_data, start=5):
    ws1.row_dimensions[idx].height = 22
    ws1.cell(row=idx, column=1, value=fid).alignment = align_center
    ws1.cell(row=idx, column=1).font = font_td_bold
    ws1.cell(row=idx, column=2, value=desc).font = font_td
    
    c_w = ws1.cell(row=idx, column=3, value=weight)
    c_w.font = font_td; c_w.alignment = align_center; c_w.number_format = '0.0'
    
    c_s = ws1.cell(row=idx, column=4, value=score)
    c_s.font = font_td_bold; c_s.alignment = align_center

    c_sub = ws1.cell(row=idx, column=5, value=f"=C{idx}*D{idx}")
    c_sub.font = font_subtotal; c_sub.fill = fill_subtotal; c_sub.alignment = align_center; c_sub.number_format = '0.0'

    c_j = ws1.cell(row=idx, column=6, value=justif)
    c_j.font = font_td; c_j.alignment = align_left

    for col_idx in range(1, 7):
        c = ws1.cell(row=idx, column=col_idx)
        c.border = thin_border
        if idx % 2 == 1 and col_idx != 5: c.fill = fill_zebra

# TCF Total (Row 18)
ws1.merge_cells("A18:D18")
c_tot1 = ws1['A18']; c_tot1.value = "SOMATÓRIO DOS FATORES (Σ TF)"; c_tot1.font = font_total_label; c_tot1.alignment = align_right
c_val1 = ws1['E18']; c_val1.value = "=SUM(E5:E17)"; c_val1.font = font_total_val; c_val1.fill = fill_subtotal; c_val1.alignment = align_center; c_val1.number_format = '0.0'
for col_idx in range(1, 7): ws1.cell(row=18, column=col_idx).border = thin_border

ws1['A20'] = "CÁLCULO TCF:"; ws1['A20'].font = font_sec
ws1['B20'] = "=0.6 + (0.01 * E18)"; ws1['B20'].font = font_badge; ws1['B20'].fill = fill_accent; ws1['B20'].alignment = align_center; ws1['B20'].number_format = '0.000'


# SHEET 2: ECF
ws2 = wb.create_sheet(title="2. Fatores Ambientais (ECF)")
ws2.views.sheetView[0].showGridLines = True

ws2['A1'] = "AVALIAÇÃO DO FATOR DE COMPLEXIDADE AMBIENTAL (ECF / FCA)"
ws2['A1'].font = font_title
ws2['A2'] = "Método de Pontos por Casos de Uso (Karner, 1993) — HRTech Core"
ws2['A2'].font = font_subtitle

ecf_headers = ["ID", "Descrição do Fator Ambiental", "Peso", "Nota (0-5)", "Subtotal (Peso × Nota)", "Justificativa Técnica no Contexto da Equipe / Projeto"]
ws2.row_dimensions[4].height = 26

for col_num, h in enumerate(ecf_headers, 1):
    c = ws2.cell(row=4, column=col_num, value=h)
    c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

ecf_data = [
    ("F1", "Processo formal de desenvolvimento de software", 1.5, 4, "Uso de metodologia formal de LPS (Linha de Produção de Software), controle de versão (Git), medição e testes."),
    ("F2", "Experiência na aplicação / domínio de negócio", 0.5, 3, "Conhecimento intermediário em rotinas de RH e folha, suportado por documento detalhado de especificação."),
    ("F3", "Experiência em Orientação a Objetos (OO)", 1.0, 4, "Boa experiência em OO, modularização JS ES6+, componentes visuais reusáveis e manipulação de estado em SPA."),
    ("F4", "Capacidade do líder analista", 0.5, 4, "Liderança com visão clara de arquitetura, elicitação de personas, storyboards e catálogo de requisitos funcionais."),
    ("F5", "Motivação da equipe", 1.0, 5, "Muito alta. Comprometimento em entregar um projeto modelo acadêmico com protótipo funcional completo e documentação ABNT."),
    ("F6", "Estabilidade dos requisitos", 2.0, 4, "Alta estabilidade. O escopo e variabilidades do HRTech Core foram bem definidos na fase de Design Thinking e especificação."),
    ("F7", "Trabalhadores em tempo parcial", -1.0, 3, "Peso negativo (-1.0). Equipe com dedicação acadêmica em tempo parcial (desfoca levemente mas mitigado por prazos)."),
    ("F8", "Dificuldade da linguagem de programação", -1.0, 2, "Peso negativo (-1.0). Linguagens padrão Web (HTML5/CSS3/JS) com baixa complexidade sintática (nota 2 de dificuldade).")
]

for idx, (fid, desc, weight, score, justif) in enumerate(ecf_data, start=5):
    ws2.row_dimensions[idx].height = 22
    ws2.cell(row=idx, column=1, value=fid).alignment = align_center
    ws2.cell(row=idx, column=1).font = font_td_bold
    ws2.cell(row=idx, column=2, value=desc).font = font_td

    c_w = ws2.cell(row=idx, column=3, value=weight)
    c_w.font = font_td; c_w.alignment = align_center; c_w.number_format = '0.0'

    c_s = ws2.cell(row=idx, column=4, value=score)
    c_s.font = font_td_bold; c_s.alignment = align_center

    c_sub = ws2.cell(row=idx, column=5, value=f"=C{idx}*D{idx}")
    c_sub.font = font_subtotal; c_sub.fill = fill_subtotal; c_sub.alignment = align_center; c_sub.number_format = '0.0'

    c_j = ws2.cell(row=idx, column=6, value=justif)
    c_j.font = font_td; c_j.alignment = align_left

    for col_idx in range(1, 7):
        c = ws2.cell(row=idx, column=col_idx)
        c.border = thin_border
        if idx % 2 == 1 and col_idx != 5: c.fill = fill_zebra

# ECF Total (Row 13)
ws2.merge_cells("A13:D13")
c_tot2 = ws2['A13']; c_tot2.value = "SOMATÓRIO DOS FATORES (Σ EF)"; c_tot2.font = font_total_label; c_tot2.alignment = align_right
c_val2 = ws2['E13']; c_val2.value = "=SUM(E5:E12)"; c_val2.font = font_total_val; c_val2.fill = fill_subtotal; c_val2.alignment = align_center; c_val2.number_format = '0.0'
for col_idx in range(1, 7): ws2.cell(row=13, column=col_idx).border = thin_border

ws2['A15'] = "CÁLCULO ECF:"; ws2['A15'].font = font_sec
ws2['B15'] = "=1.4 + (-0.03 * E13)"; ws2['B15'].font = font_badge; ws2['B15'].fill = fill_accent; ws2['B15'].alignment = align_center; ws2['B15'].number_format = '0.000'


# SHEET 3: CONSOLIDADO & CUSTOS
ws3 = wb.create_sheet(title="3. Consolidado AUCP & Custos")
ws3.views.sheetView[0].showGridLines = True

ws3['A1'] = "CONSOLIDADO UCP: AUCP, ESFORÇO, CRONOGRAMA E CUSTOS"
ws3['A1'].font = font_title
ws3['A2'] = "Método de Pontos por Casos de Uso (Karner / Schneider & Winters) — HRTech Core"
ws3['A2'].font = font_subtitle

ws3['A4'] = "1. DADOS DE ENTRADA DO PROJETO"; ws3['A4'].font = font_sec

master_inputs = [
    ("Sistema:", "HRTech Core — Sistema Modular de Gestão de RH (LPS)"),
    ("Pontos de Casos de Uso Não Ajustados (UUCP):", 178),
    ("Fator de Complexidade Técnica (TCF):", "='1. Fatores Técnicos (TCF)'!B20"),
    ("Fator de Complexidade Ambiental (ECF):", "='2. Fatores Ambientais (ECF)'!B15"),
    ("Valor da Homem-Hora (R$):", 50.00),
    ("Carga Horária Semanal por Dev (horas):", 40)
]

for idx, (lbl, val) in enumerate(master_inputs, start=5):
    ws3[f'A{idx}'] = lbl
    ws3[f'A{idx}'].font = Font(name="Arial", size=9, bold=True, color="475569")
    c_val = ws3[f'B{idx}']
    c_val.value = val
    c_val.font = font_td_bold
    ws3.row_dimensions[idx].height = 19
    if "R$" in lbl: c_val.number_format = 'R$ #,##0.00'
    elif isinstance(val, str) and "=" in val: c_val.number_format = '0.000'

# AUCP Calculation
ws3['A12'] = "2. CÁLCULO DO UCP AJUSTADO (AUCP)"; ws3['A12'].font = font_sec
ws3['A13'] = "Fórmula:"; ws3['B13'] = "AUCP = UUCP × TCF × ECF"; ws3['B13'].font = Font(name="Arial", size=10, italic=True)
ws3['A14'] = "Substituição:"; ws3['B14'] = '=CONCATENATE("AUCP = ", B6, " × ", TEXT(B7,"0.000"), " × ", TEXT(B8,"0.000"))'; ws3['B14'].font = font_td
ws3['A15'] = "AUCP RESULTANTE:"; ws3['B15'] = "=B6*B7*B8"; ws3['B15'].font = font_badge; ws3['B15'].fill = fill_accent; ws3['B15'].alignment = align_center; ws3['B15'].number_format = '0.000'

# Productivity Rules
ws3['A17'] = "3. AVALIAÇÃO DA TAXA DE PRODUTIVIDADE (REGRA DE SCHNEIDER & WINTERS)"; ws3['A17'].font = font_sec
comp_h = ["Variável", "Regra / Condição", "Quantidade Apurada", "Taxa Recomendada"]
ws3.row_dimensions[18].height = 24
for c_i, h in enumerate(comp_h, 1):
    c = ws3.cell(row=18, column=c_i, value=h); c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

ws3.cell(row=19, column=1, value="Fatores F1-F6 < 3 (X)").font = font_td_bold
ws3.cell(row=19, column=2, value="Qtd de fatores de processo/equipe com nota < 3").font = font_td
ws3.cell(row=19, column=3, value=0).font = font_td_bold; ws3.cell(row=19, column=3).alignment = align_center
ws3.cell(row=19, column=4, value="—").alignment = align_center

ws3.cell(row=20, column=1, value="Fatores F7-F8 > 3 (Y)").font = font_td_bold
ws3.cell(row=20, column=2, value="Qtd de fatores de risco com nota > 3").font = font_td
ws3.cell(row=20, column=3, value=0).font = font_td_bold; ws3.cell(row=20, column=3).alignment = align_center
ws3.cell(row=20, column=4, value="—").alignment = align_center

ws3.cell(row=21, column=1, value="Total (X + Y)").font = font_total_label
ws3.cell(row=21, column=2, value="Condição: (X + Y) <= 2 → 20 h/UCP").font = font_td_bold
ws3.cell(row=21, column=3, value="=C19+C20").font = font_total_val; ws3.cell(row=21, column=3).alignment = align_center
ws3.cell(row=21, column=4, value="20 h / UCP").font = font_total_val; ws3.cell(row=21, column=4).alignment = align_center; ws3.cell(row=21, column=4).fill = fill_subtotal

for r in range(19, 22):
    ws3.row_dimensions[r].height = 22
    for c in range(1, 5): ws3.cell(row=r, column=c).border = thin_border

# Comparative Table
ws3['A23'] = "4. TABELA COMPARATIVA DE PRODUTIVIDADE E CRONOGRAMA"; ws3['A23'].font = font_sec
t_h = ["Produtividade (h/UCP)", "Esforço Total (Horas)", "Custo Total (R$)", "1 Desenvolvedor", "2 Desenvolvedores", "4 Desenvolvedores"]
ws3.row_dimensions[24].height = 26
for c_i, h in enumerate(t_h, 1):
    c = ws3.cell(row=24, column=c_i, value=h); c.font = font_th; c.fill = fill_th; c.alignment = align_center; c.border = thin_border

scenarios = [
    (15, "15 h/UCP (Otimista)"),
    (20, "20 h/UCP (OFICIAL RECOMENDADO)"),
    (25, "25 h/UCP (Conservador)"),
    (28, "28 h/UCP (Contingência / Alto Risco)")
]

for idx, (rate, label) in enumerate(scenarios, start=25):
    ws3.row_dimensions[idx].height = 24
    ws3.cell(row=idx, column=1, value=label).font = font_td_bold if rate == 20 else font_td
    c_hrs = ws3.cell(row=idx, column=2, value=f"=$B$15*{rate}")
    c_cst = ws3.cell(row=idx, column=3, value=f"=B{idx}*$B$9")
    
    c_d1 = ws3.cell(row=idx, column=4, value=f"=B{idx}/(1*$B$10)")
    c_d2 = ws3.cell(row=idx, column=5, value=f"=B{idx}/(2*$B$10)")
    c_d4 = ws3.cell(row=idx, column=6, value=f"=B{idx}/(4*$B$10)")

    c_hrs.font = font_total_val if rate == 20 else font_td; c_hrs.number_format = '#,##0.0 "h"'
    c_cst.font = font_total_val if rate == 20 else font_td; c_cst.number_format = 'R$ #,##0.00'

    for cd in [c_d1, c_d2, c_d4]:
        cd.font = font_td; cd.number_format = '0.0 "sem"'; cd.alignment = align_center

    if rate == 20:
        for col_idx in range(1, 7): ws3.cell(row=idx, column=col_idx).fill = fill_subtotal
    for col_idx in range(1, 7): ws3.cell(row=idx, column=col_idx).border = thin_border

# Formatting column widths
for wsheet in [ws1, ws2, ws3]:
    col_w = {'A': 10, 'B': 45, 'C': 26, 'D': 20, 'E': 22, 'F': 80}
    for col_l, w in col_w.items():
        wsheet.column_dimensions[col_l].width = w

master_output = "/home/fernando/Documentos/Faculdade/Projeto de medição e analise/HRTech_Core_UCP_Completo_3_Abas.xlsx"
wb.save(master_output)
print(f"Master 3-sheet UCP Workbook saved at: {master_output}")
