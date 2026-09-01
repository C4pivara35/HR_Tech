import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build_ecf_sheet(ws):
    ws.title = "Avaliação ECF (F1-F8)"
    ws.views.sheetView[0].showGridLines = True

    COLOR_HEADER_BG = "0F172A"       # Dark Slate
    COLOR_HEADER_FG = "FFFFFF"       # White
    COLOR_ACCENT = "0D9488"          # Teal
    COLOR_ZEBRA = "F8FAFC"           # Light slate
    COLOR_SUBTOTAL_BG = "F0FDF4"     # Light green
    COLOR_SUBTOTAL_FG = "166534"     # Dark green
    COLOR_BORDER = "CBD5E1"          # Gray border

    font_title = Font(name="Arial", size=16, bold=True, color="0F172A")
    font_subtitle = Font(name="Arial", size=11, bold=True, color="0D9488")
    font_th = Font(name="Arial", size=10, bold=True, color=COLOR_HEADER_FG)
    font_td = Font(name="Arial", size=10, color="1E293B")
    font_td_bold = Font(name="Arial", size=10, bold=True, color="1E293B")
    font_subtotal = Font(name="Arial", size=10, bold=True, color=COLOR_SUBTOTAL_FG)
    font_total_label = Font(name="Arial", size=11, bold=True, color="0F172A")
    font_total_val = Font(name="Arial", size=11, bold=True, color=COLOR_SUBTOTAL_FG)
    font_ecf_result = Font(name="Arial", size=14, bold=True, color="FFFFFF")

    fill_th = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
    fill_subtotal = PatternFill(start_color=COLOR_SUBTOTAL_BG, end_color=COLOR_SUBTOTAL_BG, fill_type="solid")
    fill_ecf = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # Title Block
    ws['A1'] = "AVALIAÇÃO DO FATOR DE COMPLEXIDADE AMBIENTAL (ECF / FCA)"
    ws['A1'].font = font_title
    ws['A2'] = "Método de Pontos por Casos de Uso (Use-Case Points - UCP / Karner, 1993)"
    ws['A2'].font = font_subtitle

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    # Metadata Block
    ws['A4'] = "PROJETO:"
    ws['B4'] = "HRTech Core — Sistema Modular de Gestão de Recursos Humanos"
    ws['A5'] = "PERFIL DA EQUIPE:"
    ws['B5'] = "Estudantes de Engenharia de Software com forte conhecimento em OO, padrões de reuso (LPS) e desenvolvimento Web SPA."
    ws['A6'] = "PROCESSO E REGIME:"
    ws['B6'] = "Processo ágil (Scrum/GitFlow), requisitos estáveis definidos no documento de arquitetura e regime de trabalho acadêmico (tempo parcial)."

    for row in range(4, 7):
        ws[f'A{row}'].font = Font(name="Arial", size=9, bold=True, color="475569")
        ws[f'B{row}'].font = Font(name="Arial", size=9, color="1E293B")
        ws[f'A{row}'].alignment = align_left
        ws[f'B{row}'].alignment = align_left
        ws.row_dimensions[row].height = 18

    # Headers (Row 8)
    headers = ["ID", "Descrição do Fator Ambiental", "Peso", "Nota (0-5)", "Subtotal (Peso × Nota)", "Justificativa Técnica no Contexto da Equipe / Projeto"]
    ws.row_dimensions[8].height = 28

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = align_center
        cell.border = thin_border

    # Factors F1 to F8
    factors_data = [
        ("F1", "Processo formal de desenvolvimento de software", 1.5, 4, "Uso de metodologia formal de LPS (Linha de Produção de Software), controle de versão (Git), medição e testes."),
        ("F2", "Experiência na aplicação / domínio de negócio", 0.5, 3, "Conhecimento intermediário em rotinas de RH e folha, suportado por documento detalhado de especificação."),
        ("F3", "Experiência em Orientação a Objetos (OO)", 1.0, 4, "Boa experiência em OO, modularização JS ES6+, componentes visuais reusáveis e manipulação de estado em SPA."),
        ("F4", "Capacidade do líder analista", 0.5, 4, "Liderança com visão clara de arquitetura, elicitação de personas, storyboards e catálogo de requisitos funcionais."),
        ("F5", "Motivação da equipe", 1.0, 5, "Muito alta. Comprometimento em entregar um projeto modelo acadêmico com protótipo funcional completo e documentação ABNT."),
        ("F6", "Estabilidade dos requisitos", 2.0, 4, "Alta estabilidade. O escopo e variabilidades do HRTech Core foram bem definidos na fase de Design Thinking e especificação."),
        ("F7", "Trabalhadores em tempo parcial", -1.0, 3, "Peso negativo (-1.0). Equipe com dedicação acadêmica em tempo parcial (desfoca levemente mas mitigado por prazos)."),
        ("F8", "Dificuldade da linguagem de programação", -1.0, 2, "Peso negativo (-1.0). Linguagens padrão Web (HTML5/CSS3/JS) com baixa complexidade sintática (nota 2 de dificuldade).")
    ]

    start_row = 9
    for idx, (fid, desc, weight, score, justif) in enumerate(factors_data):
        current_row = start_row + idx
        ws.row_dimensions[current_row].height = 24

        c_id = ws.cell(row=current_row, column=1, value=fid)
        c_desc = ws.cell(row=current_row, column=2, value=desc)
        c_weight = ws.cell(row=current_row, column=3, value=weight)
        c_score = ws.cell(row=current_row, column=4, value=score)
        c_subtotal = ws.cell(row=current_row, column=5, value=f"=C{current_row}*D{current_row}")
        c_justif = ws.cell(row=current_row, column=6, value=justif)

        c_id.alignment = align_center
        c_id.font = font_td_bold

        c_desc.alignment = align_left
        c_desc.font = font_td

        c_weight.alignment = align_center
        c_weight.font = font_td
        c_weight.number_format = '0.0'

        c_score.alignment = align_center
        c_score.font = font_td_bold

        c_subtotal.alignment = align_center
        c_subtotal.font = font_subtotal
        c_subtotal.fill = fill_subtotal
        c_subtotal.number_format = '0.0'

        c_justif.alignment = align_left
        c_justif.font = font_td

        for c in [c_id, c_desc, c_weight, c_score, c_subtotal, c_justif]:
            c.border = thin_border
            if idx % 2 == 1 and c != c_subtotal:
                c.fill = fill_zebra

    # Total Row (Row 17)
    total_row = start_row + len(factors_data)
    ws.row_dimensions[total_row].height = 26

    c_tot_label = ws.cell(row=total_row, column=1, value="SOMATÓRIO DOS FATORES (Σ EF)")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    c_tot_label.font = font_total_label
    c_tot_label.alignment = align_right

    c_tot_val = ws.cell(row=total_row, column=5, value=f"=SUM(E9:E{total_row-1})")
    c_tot_val.font = font_total_val
    c_tot_val.fill = fill_subtotal
    c_tot_val.alignment = align_center
    c_tot_val.number_format = '0.0'

    for col in range(1, 7):
        ws.cell(row=total_row, column=col).border = thin_border

    # Calculation Block
    ws.cell(row=19, column=1, value="MEMÓRIA DE CÁLCULO DO ECF:").font = Font(name="Arial", size=11, bold=True, color="0F172A")

    ws.cell(row=20, column=1, value="Fórmula Oficial:").font = Font(name="Arial", size=10, bold=True, color="475569")
    ws.cell(row=20, column=2, value="ECF = 1,4 + (-0,03 × Σ EF)").font = Font(name="Arial", size=10, italic=True, color="1E293B")

    ws.cell(row=21, column=1, value="Substituição:").font = Font(name="Arial", size=10, bold=True, color="475569")
    ws.cell(row=21, column=2, value=f"=CONCATENATE(\"ECF = 1,4 + (-0,03 × \", TEXT(E{total_row}, \"0.0\"), \")\")").font = Font(name="Arial", size=10, color="1E293B")

    ws.cell(row=23, column=1, value="VALOR FINAL DE ECF:").font = Font(name="Arial", size=12, bold=True, color="0F172A")
    c_ecf = ws.cell(row=23, column=2, value=f"=1.4 + (-0.03 * E{total_row})")
    c_ecf.font = font_ecf_result
    c_ecf.fill = fill_ecf
    c_ecf.alignment = align_center
    c_ecf.number_format = '0.000'
    ws.row_dimensions[23].height = 30

    column_widths = {
        'A': 8,
        'B': 45,
        'C': 10,
        'D': 12,
        'E': 24,
        'F': 85
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

# Create standalone ECF workbook
wb_ecf = openpyxl.Workbook()
ws_ecf = wb_ecf.active
build_ecf_sheet(ws_ecf)
output_ecf = "/home/fernando/Documentos/Faculdade/Projeto de medição e analise/Avaliacao_ECF_UCP_HRTech_Core.xlsx"
wb_ecf.save(output_ecf)
print(f"Standalone ECF Excel created at: {output_ecf}")

# Also update a combined workbook containing both TCF (Sheet 1) and ECF (Sheet 2)
from create_tcf_excel import wb as wb_tcf
ws_tcf = wb_tcf.active
ws_tcf.title = "Avaliação TCF (T1-T13)"
ws2 = wb_tcf.create_sheet(title="Avaliação ECF (F1-F8)")
build_ecf_sheet(ws2)

output_combined = "/home/fernando/Documentos/Faculdade/Projeto de medição e analise/Avaliacao_TCF_E_ECF_UCP_HRTech_Core.xlsx"
wb_tcf.save(output_combined)
print(f"Combined TCF + ECF Excel created at: {output_combined}")
