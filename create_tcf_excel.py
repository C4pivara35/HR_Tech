import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Avaliação TCF (UCP)"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Colors
COLOR_HEADER_BG = "0F172A"       # Dark Slate
COLOR_HEADER_FG = "FFFFFF"       # White
COLOR_ACCENT = "0D9488"          # Teal
COLOR_ZEBRA = "F8FAFC"           # Light slate
COLOR_SUBTOTAL_BG = "F0FDF4"     # Light green
COLOR_SUBTOTAL_FG = "166534"     # Dark green
COLOR_BORDER = "CBD5E1"          # Gray border

# Styles
font_title = Font(name="Arial", size=16, bold=True, color="0F172A")
font_subtitle = Font(name="Arial", size=11, bold=True, color="0D9488")
font_section = Font(name="Arial", size=12, bold=True, color="0F172A")
font_th = Font(name="Arial", size=10, bold=True, color=COLOR_HEADER_FG)
font_td = Font(name="Arial", size=10, color="1E293B")
font_td_bold = Font(name="Arial", size=10, bold=True, color="1E293B")
font_subtotal = Font(name="Arial", size=10, bold=True, color=COLOR_SUBTOTAL_FG)
font_total_label = Font(name="Arial", size=11, bold=True, color="0F172A")
font_total_val = Font(name="Arial", size=11, bold=True, color=COLOR_SUBTOTAL_FG)
font_tcf_result = Font(name="Arial", size=14, bold=True, color="FFFFFF")

fill_th = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
fill_subtotal = PatternFill(start_color=COLOR_SUBTOTAL_BG, end_color=COLOR_SUBTOTAL_BG, fill_type="solid")
fill_meta = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_tcf = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# Title Block
ws['A1'] = "AVALIAÇÃO DO FATOR DE COMPLEXIDADE TÉCNICA (TCF)"
ws['A1'].font = font_title
ws['A2'] = "Método de Pontos por Casos de Uso (Use-Case Points - UCP / Karner, 1993)"
ws['A2'].font = font_subtitle

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 20

# Metadata Block
ws['A4'] = "PROJETO:"
ws['B4'] = "HRTech Core — Sistema Modular de Gestão de Recursos Humanos"
ws['A5'] = "APLICAÇÃO:"
ws['B5'] = "Plataforma SaaS modular baseada em Linha de Produção de Software (LPS) com variabilidade para Tecnologia, Indústria e Financeiro/Seguros."
ws['A6'] = "ARQUITETURA:"
ws['B6'] = "Web SPA (HTML5/CSS3/JS ES6+), RESTful APIs, banco relacional multi-tenant e controle por feature toggles."

for row in range(4, 7):
    ws[f'A{row}'].font = Font(name="Arial", size=9, bold=True, color="475569")
    ws[f'B{row}'].font = Font(name="Arial", size=9, color="1E293B")
    ws[f'A{row}'].alignment = align_left
    ws[f'B{row}'].alignment = align_left
    ws.row_dimensions[row].height = 18

# Table Headers (Row 8)
headers = ["ID", "Descrição do Fator", "Peso", "Nota (0-5)", "Subtotal (Peso × Nota)", "Justificativa Técnica no Contexto do Projeto"]
ws.row_dimensions[8].height = 28

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col_num, value=header)
    cell.font = font_th
    cell.fill = fill_th
    cell.alignment = align_center
    cell.border = thin_border

# Data Rows (T1 to T13)
factors_data = [
    ("T1", "Sistemas distribuídos", 2.0, 4, "Arquitetura Web SPA cliente-servidor distribuída com APIs REST e multi-tenancy em nuvem."),
    ("T2", "Objetivos de desempenho e tempo de resposta", 1.0, 4, "Crítico para registros de ponto em horários de pico (resposta esperada < 200 ms — RNF02)."),
    ("T3", "Eficiência do usuário final (on-line)", 1.0, 4, "Navegação SPA fluida, painéis resumidos e aprovações em 1 clique para agilizar a operação."),
    ("T4", "Complexidade interna de processamento", 1.0, 3, "Algoritmos de cálculo de saldo de banco de horas, escalas 12x36, abonos e dupla custódia."),
    ("T5", "Código deve ser reutilizado", 1.0, 5, "Essencial (LPS). Meta de reuso > 75% com 6 ativos reutilizáveis documentados (AR01 a AR06)."),
    ("T6", "Facilidade de instalação", 0.5, 3, "SaaS em nuvem sem instalação no cliente, mas com setup automático de novos tenants em < 8h."),
    ("T7", "Facilidade de uso (Usabilidade)", 0.5, 4, "Atende desde diretores de RH até operadores de fábrica, exigindo alta usabilidade e clareza."),
    ("T8", "Portabilidade", 2.0, 4, "Suporte a múltiplos navegadores em desktops, tablets industriais e smartphones."),
    ("T9", "Facilidade de mudança (Manutenibilidade)", 1.0, 5, "Essencial. Arquitetura modular desacoplada para ativação de regras e módulos por tenant."),
    ("T10", "Concorrência", 1.0, 4, "Elevado número de acessos simultâneos ao bater ponto nas trocas de turno e ao consultar dashboards."),
    ("T11", "Características especiais de segurança", 1.0, 4, "Controle RBAC, trilha de auditoria imutável, conformidade com a LGPD e dupla custódia."),
    ("T12", "Acesso direto para terceiros", 1.0, 3, "Integração API com cartões de benefícios (Caju/Flash), eSocial e Portal do Corretor parceiro."),
    ("T13", "Treinamento especial dos usuários", 1.0, 2, "Interface autoexplicativa com design intuitivo para dispensar treinamentos complexos.")
]

start_row = 9
for idx, (fid, desc, weight, score, justif) in enumerate(factors_data):
    current_row = start_row + idx
    ws.row_dimensions[current_row].height = 24

    c_id = ws.cell(row=current_row, column=1, value=fid)
    c_desc = ws.cell(row=current_row, column=2, value=desc)
    c_weight = ws.cell(row=current_row, column=3, value=weight)
    c_score = ws.cell(row=current_row, column=4, value=score)
    # Excel Formula for Subtotal
    c_subtotal = ws.cell(row=current_row, column=5, value=f"=C{current_row}*D{current_row}")
    c_justif = ws.cell(row=current_row, column=6, value=justif)

    # Formatting
    c_id.alignment = align_center
    c_id.font = font_td_bold
    
    c_desc.alignment = align_left
    c_desc.font = font_td

    c_weight.alignment = align_center
    c_weight.font = font_td
    c_weight.number_format = '0.0'

    c_score.alignment = align_center
    c_score.font = font_td_bold

    c_subtotal.alignment = align_center
    c_subtotal.font = font_subtotal
    c_subtotal.fill = fill_subtotal
    c_subtotal.number_format = '0.0'

    c_justif.alignment = align_left
    c_justif.font = font_td

    # Apply borders & zebra fill
    for c in [c_id, c_desc, c_weight, c_score, c_subtotal, c_justif]:
        c.border = thin_border
        if idx % 2 == 1 and c != c_subtotal:
            c.fill = fill_zebra

# Total Row (Row 22)
total_row = start_row + len(factors_data)
ws.row_dimensions[total_row].height = 26

c_tot_label = ws.cell(row=total_row, column=1, value="SOMATÓRIO DOS FATORES (Σ TF)")
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
c_tot_label.font = font_total_label
c_tot_label.alignment = align_right

c_tot_val = ws.cell(row=total_row, column=5, value=f"=SUM(E9:E{total_row-1})")
c_tot_val.font = font_total_val
c_tot_val.fill = fill_subtotal
c_tot_val.alignment = align_center
c_tot_val.number_format = '0.0'

for col in range(1, 7):
    ws.cell(row=total_row, column=col).border = thin_border

# Calculation & Result Block (Rows 24 to 28)
ws.cell(row=24, column=1, value="MEMÓRIA DE CÁLCULO DO TCF:").font = Font(name="Arial", size=11, bold=True, color="0F172A")

ws.cell(row=25, column=1, value="Fórmula:").font = Font(name="Arial", size=10, bold=True, color="475569")
ws.cell(row=25, column=2, value="TCF = 0,6 + (0,01 × Σ TF)").font = Font(name="Arial", size=10, italic=True, color="1E293B")

ws.cell(row=26, column=1, value="Substituição:").font = Font(name="Arial", size=10, bold=True, color="475569")
ws.cell(row=26, column=2, value=f"=CONCATENATE(\"TCF = 0,6 + (0,01 × \", TEXT(E{total_row}, \"0.0\"), \")\")").font = Font(name="Arial", size=10, color="1E293B")

# TCF Final Result Cell
ws.cell(row=28, column=1, value="VALOR FINAL DE TCF:").font = Font(name="Arial", size=12, bold=True, color="0F172A")
c_tcf = ws.cell(row=28, column=2, value=f"=0.6 + (0.01 * E{total_row})")
c_tcf.font = font_tcf_result
c_tcf.fill = fill_tcf
c_tcf.alignment = align_center
c_tcf.number_format = '0.000'
ws.row_dimensions[28].height = 30

# Column Widths
column_widths = {
    'A': 8,
    'B': 42,
    'C': 10,
    'D': 12,
    'E': 24,
    'F': 85
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save Excel file
output_path = "/home/fernando/Documentos/Faculdade/Projeto de medição e analise/Avaliacao_TCF_UCP_HRTech_Core.xlsx"
wb.save(output_path)
print(f"Excel file successfully generated at: {output_path}")
